import requests
import datetime as dt

import apimoex
import pandas as pd

from openpyxl import load_workbook
import numpy as np

import smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def get_data(currencies=[]):
    # set arguments for request
	date_till = dt.datetime.now(dt.timezone(dt.timedelta(hours=3))).date()
	date_from = date_till - dt.timedelta(days=30)

	request_urls = [('https://iss.moex.com/iss/statistics/engines/futures/'
					 'markets/indicativerates/securities/{}.json'
					 '?till={}&from={}'.format(currency, str(date_till), str(date_from)))
										   for currency in currencies]
	dfs = [] # list of currency dataframes
	with requests.Session() as session:
		for req in request_urls:
			iss = apimoex.ISSClient(session, req)
			data = iss.get()
			dfs.append(pd.DataFrame(data['securities']))
	return dfs


def create_excel(dfs=[], descending=True):
	output_df = pd.DataFrame()
	for df in dfs:
		if descending:
			df = df[::-1]
		cur_name = df.secid[0]
		output_df[f'Дата {cur_name}'] = pd.to_datetime(df.tradedate, format='%Y-%m-%d').dt.date
		output_df[f'Курс {cur_name}'] = df.rate
		output_df[f'Изменение {cur_name}'] = output_df[f'Курс {cur_name}'].diff(periods=-1)
	output_df['Отношение EUR к USD'] = output_df['Курс EUR/RUB'] / output_df['Курс USD/RUB']

	date_from = output_df['Дата USD/RUB'].max()
	file_name = f'Currencies_{str(date_from)}.xlsx'
	output_df.to_excel(file_name, index=False, engine='openpyxl')
	num_strings = output_df.shape[0]
	return (file_name, num_strings)


def autowidth(ws):
	column_widths = {}
	for col in ws.columns:
		col_lens = np.array([len(str(cell.value)) for cell in col])
		column_widths[col[0].column] = col_lens.max() + 2

	for col, value in column_widths.items():
		try:
			ws.column_dimensions[col].width = value
		except (TypeError, ValueError) as ex:
			column_names = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G'}
			ws.column_dimensions[column_names[col]].width = value


def formatting(ws):
	for i, row in enumerate(ws.rows):
		if i == 0:
			continue
		n_row = i + 1
		for col in [1, 4]:
			ws.cell(n_row, col).number_format = 'dd.mm.yyyy'
		for col in [2, 3, 5, 6]:
			ws.cell(n_row, col).number_format = '#,##0.00 [$р.-419];-#,##0.00 [$р.-419]' # '_-* #,##0.00 ₽_-'
		ws.cell(n_row, 7).number_format = '0.0000'


def send_email(remote_server, port,
			   email_from, email_to,
			   subject, file_name, num_strings):
	pwd = input('Your password: ')

	message = MIMEMultipart()
	message['Name'] = 'Max'
	message['From'] = email_from
	message['To'] = email_to
	message['Subject'] = subject

	# create body message
	strings = ''
	if 11 <= num_strings <= 14:
		strings = 'строк'
	elif num_strings % 10 == 1:
		strings = 'строку'
	elif num_strings % 10 in [2, 3, 4]:
		strings = 'строки'
	else:
		strings = 'строк'
	body = ('Здравствуйте!\nВо вложении файл с курсами валют за последний месяц. '
			'Файл содержит {} {} данных.\n\n'
			'P.S. Письмо отправлено с помощью Python.\n'.format(num_strings, strings))
	message.attach(MIMEText(body, 'plain'))

	# add attachment
	with open(file_name, 'rb') as att:
		part = MIMEBase('application', 'octet-stream')
		part.set_payload(att.read())
	encoders.encode_base64(part)
	part.add_header('Content-Disposition',
					'attachment', filename=file_name)
	message.attach(part)

	# final text
	text = message.as_string()

	# send email
	context = ssl.create_default_context()
	with smtplib.SMTP_SSL(remote_server, port, context=context) as server:
		server.login(email_from, pwd)
		server.sendmail(email_from, email_to, text)


if __name__ == '__main__':
	currencies = ['USD/RUB', 'EUR/RUB']

	remote_server = 'smtp.gmail.com'
	port = 465
	email_from = '***@gmail.com'
	email_to = '***@gmail.com'
	subject = 'Курсы валют ' + ', '.join(currencies)

	data = get_data(currencies)
	file_name, n = create_excel(data)

	wb = load_workbook(file_name)
	ws = wb['Sheet1']
	autowidth(ws)
	formatting(ws)
	wb.save(file_name)

	send_email(remote_server, port,
			   email_from, email_to,
			   subject, file_name, n)
